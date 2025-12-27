# -*- coding: utf-8 -*-
"""
Interface des paramètres
"""
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
                             QComboBox, QFrame, QMessageBox, QHeaderView, QTabWidget,
                             QFormLayout, QGroupBox, QCheckBox, QSpinBox, QFileDialog)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPixmap
from core.auth import auth_manager
from core.logger import logger
from database.db_manager import db
import config
import openpyxl
import os
from ui.permission_dialog import PermissionDialog

class SettingsPage(QWidget):
    """Page de configuration"""
    
    # Signal pour changement de thème (si implémenté dynamiquement)
    theme_changed = pyqtSignal(bool) # True = Dark mode
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.load_users()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # En-tête avec gradient
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #475569, stop:1 #334155);
                border-radius: 12px;
                padding: 20px;
                margin-bottom: 20px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        
        title_layout = QVBoxLayout()
        header = QLabel("⚙️ Paramètres")
        header.setStyleSheet("font-size: 24px; font-weight: bold; color: white; background: transparent;")
        title_layout.addWidget(header)
        
        subtitle = QLabel("Configuration générale, utilisateurs et sauvegardes")
        subtitle.setStyleSheet("font-size: 14px; color: rgba(255,255,255,0.9); background: transparent;")
        title_layout.addWidget(subtitle)
        
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        layout.addWidget(header_frame)
        
        # Onglets stylisés
        tabs = QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                background: white;
                border-radius: 12px;
                top: -1px; 
            }
            QTabBar::tab {
                background: #f1f5f9;
                border: 1px solid #e2e8f0;
                padding: 12px 20px;
                margin-right: 5px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                color: #64748b;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom-color: white;
                color: #0f172a;
            }
            QTabBar::tab:hover {
                background: #e2e8f0;
            }
        """)
        
        is_admin = auth_manager.is_admin()
        
        if is_admin:
            # Onglet Utilisateurs
            self.users_tab = self.create_users_tab()
            tabs.addTab(self.users_tab, "👥 Utilisateurs")
            
            # Onglet Données (Export)
            self.data_tab = self.create_data_tab()
            tabs.addTab(self.data_tab, "💾 Données")

            # Onglet Magasin
            self.store_tab = self.create_store_tab()
            tabs.addTab(self.store_tab, "🏪 Magasin")
        
        # Onglet Tutoriel (Pour tous)
        self.tutorial_tab = self.create_tutorial_tab()
        tabs.addTab(self.tutorial_tab, "📖 Tutoriel")
        
        # Onglet À propos (Pour tous)
        self.about_tab = self.create_about_tab()
        tabs.addTab(self.about_tab, "ℹ️ À propos")
        
        layout.addWidget(tabs)
        self.setLayout(layout)

    def create_data_tab(self):
        """Onglet de gestion des données"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Section Configuration Sauvegarde Auto
        backup_config_group = QGroupBox("⚙️ Configuration Sauvegarde Automatique")
        backup_form = QFormLayout()
        
        self.auto_backup_check = QCheckBox("Activer la sauvegarde automatique")
        self.auto_backup_check.setChecked(True) # Valeur par défaut, sera écrasée par la DB
        
        self.backup_interval_spin = QSpinBox()
        self.backup_interval_spin.setRange(1, 48)
        self.backup_interval_spin.setSuffix(" heures")
        self.backup_interval_spin.setValue(config.BACKUP_CONFIG.get("backup_interval_hours", 5))
        
        # Charger les valeurs depuis la DB
        try:
            # Check enabled
            res_enabled = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'auto_backup_enabled'")
            if res_enabled:
                self.auto_backup_check.setChecked(res_enabled['setting_value'] == '1')
            else:
                self.auto_backup_check.setChecked(config.BACKUP_CONFIG.get("auto_backup", True))
                
            # Check interval
            res_interval = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'backup_interval_hours'")
            if res_interval:
                self.backup_interval_spin.setValue(int(res_interval['setting_value']))
        except Exception as e:
            logger.error(f"Erreur chargement config backup: {e}")
            
        backup_form.addRow(self.auto_backup_check)
        backup_form.addRow("Intervalle:", self.backup_interval_spin)
        
        save_backup_btn = QPushButton("💾 Enregistrer la configuration")
        save_backup_btn.clicked.connect(self.save_backup_config)
        save_backup_btn.setStyleSheet("background-color: #34495e; color: white;")
        backup_form.addRow(save_backup_btn)
        
        backup_config_group.setLayout(backup_form)
        layout.addWidget(backup_config_group)
        
        # Section Export
        export_group = QGroupBox("📤 Sauvegarde / Exportation")
        export_form = QFormLayout()
        
        export_info = QLabel("Créez une sauvegarde complète de toutes vos données en cas de problème PC ou logiciel.")
        export_info.setStyleSheet("color: gray;")
        export_info.setWordWrap(True)
        export_form.addRow(export_info)
        
        export_btn = QPushButton("💾 Créer une Sauvegarde Complète (Excel)")
        export_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 12px; font-weight: bold; font-size: 14px;")
        export_btn.clicked.connect(self.export_data)
        export_form.addRow(export_btn)
        
        export_group.setLayout(export_form)
        layout.addWidget(export_group)
        
        # Section Import
        import_group = QGroupBox("📥 Restauration / Importation")
        import_form = QFormLayout()
        
        import_info = QLabel("Restaurez vos données depuis un fichier de sauvegarde Excel précédemment créé.")
        import_info.setStyleSheet("color: gray;")
        import_info.setWordWrap(True)
        import_form.addRow(import_info)
        
        import_btn = QPushButton("📂 Restaurer depuis une Sauvegarde")
        import_btn.setStyleSheet("background-color: #3498db; color: white; padding: 12px; font-weight: bold; font-size: 14px;")
        import_btn.clicked.connect(self.import_data)
        import_form.addRow(import_btn)
        
        import_group.setLayout(import_form)
        layout.addWidget(import_group)
        
        # Section Réinitialisation (DANGER)
        reset_group = QGroupBox("⚠️ Zone Danger - Réinitialisation")
        reset_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #e74c3c;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                color: #e74c3c;
            }
        """)
        reset_form = QFormLayout()
        
        reset_info = QLabel("⚠️ ATTENTION: Cette action supprimera TOUTES les données (produits, ventes, clients, fournisseurs). Cette action est IRRÉVERSIBLE!")
        reset_info.setStyleSheet("color: #e74c3c; font-weight: bold;")
        reset_info.setWordWrap(True)
        reset_form.addRow(reset_info)
        
        reset_btn = QPushButton("🗑️ RÉINITIALISER TOUTES LES DONNÉES")
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c; 
                color: white; 
                padding: 12px; 
                font-weight: bold; 
                font-size: 14px;
                border: 2px solid #c0392b;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        reset_btn.clicked.connect(self.reset_all_data)
        reset_form.addRow(reset_btn)
        
        reset_group.setLayout(reset_form)
        layout.addWidget(reset_group)
        
        layout.addStretch()
        tab.setLayout(layout)
        return tab

    def save_backup_config(self):
        """Enregistrer la configuration de sauvegarde"""
        try:
            enabled = '1' if self.auto_backup_check.isChecked() else '0'
            interval = str(self.backup_interval_spin.value())
            
            # Upsert settings
            db.execute_update("INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES ('auto_backup_enabled', ?)", (enabled,))
            db.execute_update("INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES ('backup_interval_hours', ?)", (interval,))
            
            QMessageBox.information(self, "Succès", "Configuration de sauvegarde enregistrée.\nRedémarrez l'application pour appliquer les changements.")
            logger.info(f"Config backup mise à jour: Enabled={enabled}, Interval={interval}h")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'enregistrement: {e}")
            logger.error(f"Erreur save backup config: {e}")

    def export_data(self):
        """Exporter les données en Excel (sauvegarde complète)"""
        try:
            from datetime import datetime
            default_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(self, "Sauvegarder les données", 
                                                    str(config.DATA_DIR / default_name), 
                                                    "Fichiers Excel (*.xlsx)")
            if not filename:
                return

            wb = openpyxl.Workbook()
            
            # 1. Produits
            ws_prod = wb.active
            ws_prod.title = "Produits"
            products = db.execute_query("SELECT barcode, name, category, purchase_price, selling_price, stock_quantity, min_stock FROM products")
            ws_prod.append(["Code-barres", "Nom", "Catégorie", "PA", "PV", "Stock", "Stock Min"])
            for p in products:
                ws_prod.append([p['barcode'], p['name'], p.get('category', ''), p['purchase_price'], p['selling_price'], p['stock_quantity'], p.get('min_stock', 0)])
                
            # 2. Ventes
            ws_sales = wb.create_sheet("Ventes")
            sales = db.execute_query("SELECT sale_number, total_amount, payment_method, sale_date, customer_id FROM sales")
            ws_sales.append(["N° Vente", "Montant Total", "Paiement", "Date", "Client ID"])
            for s in sales:
                ws_sales.append([s['sale_number'], s['total_amount'], s['payment_method'], s['sale_date'], s.get('customer_id', '')])

            # 3. Détails Ventes
            ws_items = wb.create_sheet("Details_Ventes")
            items = db.execute_query("SELECT sale_id, product_id, quantity, unit_price, total FROM sale_items")
            ws_items.append(["ID Vente", "ID Produit", "Quantité", "Prix Unitaire", "Total"])
            for i in items:
                ws_items.append([i['sale_id'], i['product_id'], i['quantity'], i['unit_price'], i['total']])

            # 4. Clients
            ws_cust = wb.create_sheet("Clients")
            customers = db.execute_query("SELECT full_name, phone, current_credit, total_purchases FROM customers")
            ws_cust.append(["Nom", "Téléphone", "Dette", "Total Achats"])
            for c in customers:
                ws_cust.append([c['full_name'], c['phone'], c['current_credit'], c['total_purchases']])

            # 5. Fournisseurs
            ws_sup = wb.create_sheet("Fournisseurs")
            try:
                suppliers = db.execute_query("SELECT name, phone, email, address FROM suppliers")
                ws_sup.append(["Nom", "Téléphone", "Email", "Adresse"])
                for sup in suppliers:
                    ws_sup.append([sup['name'], sup.get('phone', ''), sup.get('email', ''), sup.get('address', '')])
            except:
                ws_sup.append(["Aucune donnée fournisseur"])
            
            wb.save(filename)
            logger.info(f"Sauvegarde créée: {filename}")
            QMessageBox.information(self, "✅ Succès", f"Sauvegarde complète créée avec succès!\n\nFichier: {filename}\n\nDonnées incluses:\n• Produits\n• Ventes et détails\n• Clients\n• Fournisseurs")
            
        except Exception as e:
            logger.error(f"Erreur export excel: {e}")
            QMessageBox.critical(self, "Erreur", f"Échec de l'exportation: {e}")

    def import_data(self):
        """Importer les données depuis une sauvegarde Excel"""
        reply = QMessageBox.warning(self, "⚠️ Attention", 
            "L'importation va REMPLACER les données existantes.\n\nÊtes-vous sûr de vouloir continuer?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply != QMessageBox.Yes:
            return
            
        try:
            filename, _ = QFileDialog.getOpenFileName(self, "Sélectionner le fichier de sauvegarde",
                                                     str(config.DATA_DIR),
                                                     "Fichiers Excel (*.xlsx)")
            if not filename:
                return

            wb = openpyxl.load_workbook(filename)
            imported_counts = {}
            
            # 1. Importer Produits
            if "Produits" in wb.sheetnames:
                ws = wb["Produits"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # barcode exists
                        try:
                            db.execute_update("""
                                INSERT OR REPLACE INTO products (barcode, name, category, purchase_price, selling_price, stock_quantity, min_stock)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (row[0], row[1], row[2] or '', row[3] or 0, row[4] or 0, row[5] or 0, row[6] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Produits"] = count

            # 2. Importer Clients
            if "Clients" in wb.sheetnames:
                ws = wb["Clients"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # nom exists
                        try:
                            db.execute_update("""
                                INSERT OR IGNORE INTO customers (full_name, phone, current_credit, total_purchases)
                                VALUES (?, ?, ?, ?)
                            """, (row[0], row[1] or '', row[2] or 0, row[3] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Clients"] = count

            # 3. Importer Fournisseurs
            if "Fournisseurs" in wb.sheetnames:
                ws = wb["Fournisseurs"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0] != "Aucune donnée fournisseur":
                        try:
                            db.execute_update("""
                                INSERT OR IGNORE INTO suppliers (name, phone, email, address)
                                VALUES (?, ?, ?, ?)
                            """, (row[0], row[1] or '', row[2] or '', row[3] or ''))
                            count += 1
                        except:
                            pass
                imported_counts["Fournisseurs"] = count
            
            # Résumé
            summary = "\n".join([f"• {k}: {v} enregistrements" for k, v in imported_counts.items()])
            logger.info(f"Restauration depuis: {filename}")
            QMessageBox.information(self, "✅ Restauration Terminée", 
                f"Données restaurées avec succès!\n\n{summary}")
            
        except Exception as e:
            logger.error(f"Erreur import excel: {e}")
            QMessageBox.critical(self, "Erreur", f"Échec de la restauration: {e}")

    def reset_all_data(self):
        """Réinitialiser toutes les données de l'application"""
        # Première confirmation
        reply1 = QMessageBox.critical(self, "⚠️ ATTENTION - Réinitialisation", 
            "Vous êtes sur le point de SUPPRIMER DÉFINITIVEMENT toutes les données:\n\n"
            "• Tous les produits\n"
            "• Toutes les ventes et détails\n"
            "• Tous les clients\n"
            "• Tous les fournisseurs\n"
            "• Toutes les transactions\n\n"
            "⚠️ CETTE ACTION EST IRRÉVERSIBLE!\n\n"
            "Voulez-vous vraiment continuer?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply1 != QMessageBox.Yes:
            return
        
        # Deuxième confirmation avec mot de passe
        from PyQt5.QtWidgets import QInputDialog
        password, ok = QInputDialog.getText(self, "Confirmation par mot de passe",
            "Tapez votre mot de passe pour confirmer la réinitialisation:",
            QLineEdit.Password)
        
        if not ok or not password:
            return
        
        # Vérifier le mot de passe
        current_user = auth_manager.get_current_user()
        if not current_user:
            QMessageBox.critical(self, "Erreur", "Utilisateur non connecté")
            return
        
        success, _, _ = auth_manager.login(current_user['username'], password)
        if not success:
            QMessageBox.critical(self, "Erreur", "Mot de passe incorrect!")
            return
        
        try:
            # Créer une sauvegarde avant suppression
            from datetime import datetime
            from core.backup import backup_manager
            backup_manager.create_backup()
            
            # Supprimer toutes les données (garder les utilisateurs et catégories)
            tables_to_clear = [
                'sale_items',
                'sales',
                'customer_credit_transactions',
                'supplier_transactions',
                'price_history',
                'products',
                'customers',
                'suppliers',
                'audit_log'
            ]
            
            for table in tables_to_clear:
                try:
                    db.execute_update(f"DELETE FROM {table}")
                except Exception as e:
                    logger.error(f"Erreur suppression {table}: {e}")
            
            logger.info("⚠️ RÉINITIALISATION COMPLÈTE effectuée par l'utilisateur")
            QMessageBox.information(self, "✅ Réinitialisation Terminée", 
                "Toutes les données ont été supprimées.\n\n"
                "Une sauvegarde a été créée avant la suppression dans le dossier 'backups'.\n\n"
                "Redémarrez l'application pour un nouvel état propre.")
            
        except Exception as e:
            logger.error(f"Erreur réinitialisation: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la réinitialisation: {e}")

    # ... (rest of methods)
        
    def create_users_tab(self):
        """Onglet de gestion des utilisateurs"""
        tab = QWidget()
        layout = QHBoxLayout()
        
        # Liste des utilisateurs (Gauche)
        list_layout = QVBoxLayout()
        list_layout.addWidget(QLabel("<b>Liste des utilisateurs</b>"))
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)  # 4 colonnes avec Actions
        self.users_table.setHorizontalHeaderLabels(["Utilisateur", "Nom", "Rôle", "Actions"])
        self.users_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.users_table.setColumnWidth(3, 100)  # Largeur fixe pour Actions
        list_layout.addWidget(self.users_table)
        
        refresh_btn = QPushButton("🔄 Actualiser")
        refresh_btn.clicked.connect(self.load_users)
        list_layout.addWidget(refresh_btn)
        
        layout.addLayout(list_layout, 2)
        
        # Formulaire d'ajout (Droite)
        form_group = QGroupBox("Ajouter un utilisateur")
        form_layout = QFormLayout()
        
        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.fullname_input = QLineEdit()
        self.role_input = QComboBox()
        self.role_input.addItem("Caissier", "cashier")
        self.role_input.addItem("Administrateur", "admin")
        
        form_layout.addRow("Nom d'utilisateur:", self.username_input)
        form_layout.addRow("Mot de passe:", self.password_input)
        form_layout.addRow("Nom complet:", self.fullname_input)
        form_layout.addRow("Rôle:", self.role_input)
        
        add_btn = QPushButton("➕ Créer l'utilisateur")
        add_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 10px; font-weight: bold;")
        add_btn.clicked.connect(self.add_user)
        form_layout.addRow(add_btn)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group, 1)
        
        tab.setLayout(layout)
        return tab
        
    def create_appearance_tab(self):
        """Onglet apparence"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Message indiquant que le thème est fixe
        info_label = QLabel("✅ L'application utilise le thème clair par défaut.")
        info_label.setStyleSheet("font-size: 14px; color: #7f8c8d; padding: 20px;")
        layout.addWidget(info_label)
        
        layout.addStretch()
        
        tab.setLayout(layout)
        return tab
        
    def create_store_tab(self):
        """Onglet infos magasin"""
        tab = QWidget()
        layout = QFormLayout()
        
        # Charger config actuelle
        store_config = config.STORE_CONFIG
        
        self.store_name = QLineEdit(store_config.get('name', ''))
        self.store_phone = QLineEdit(store_config.get('phone', ''))
        self.store_address = QLineEdit(store_config.get('address', ''))
        self.store_city = QLineEdit(store_config.get('city', ''))
        
        # Champs fiscaux
        self.store_nif = QLineEdit(store_config.get('tax_id', ''))  # NIF
        self.store_nis = QLineEdit(store_config.get('nis', ''))      # NIS
        self.store_rc = QLineEdit(store_config.get('rc', ''))        # RC
        self.store_ai = QLineEdit(store_config.get('ai', ''))        # AI
        
        save_btn = QPushButton("💾 Sauvegarder")
        save_btn.setDefault(True)
        save_btn.setAutoDefault(True)
        save_btn.clicked.connect(self.save_store_settings)
        
        layout.addRow("Nom du magasin:", self.store_name)
        layout.addRow("Téléphone:", self.store_phone)
        layout.addRow("Adresse:", self.store_address)
        layout.addRow("Ville:", self.store_city)
        layout.addRow("NIF:", self.store_nif)
        layout.addRow("NIS:", self.store_nis)
        layout.addRow("RC:", self.store_rc)
        layout.addRow("AI:", self.store_ai)
        layout.addRow(save_btn)
        
        tab.setLayout(layout)
        return tab

    def save_store_settings(self):
        """Sauvegarder les paramètres du magasin"""
        try:
            # Mettre à jour la config en mémoire
            config.STORE_CONFIG['name'] = self.store_name.text()
            config.STORE_CONFIG['phone'] = self.store_phone.text()
            config.STORE_CONFIG['address'] = self.store_address.text()
            config.STORE_CONFIG['city'] = self.store_city.text()
            config.STORE_CONFIG['tax_id'] = self.store_nif.text()
            config.STORE_CONFIG['nis'] = self.store_nis.text()
            config.STORE_CONFIG['rc'] = self.store_rc.text()
            config.STORE_CONFIG['ai'] = self.store_ai.text()
            
            # Sauvegarder en base de données
            # Table settings avec (setting_key, setting_value)
            settings_to_save = {
                'store_name': config.STORE_CONFIG['name'],
                'store_phone': config.STORE_CONFIG['phone'],
                'store_address': config.STORE_CONFIG['address'],
                'store_city': config.STORE_CONFIG['city'],
                'store_nif': config.STORE_CONFIG['tax_id'],
                'store_nis': config.STORE_CONFIG['nis'],
                'store_rc': config.STORE_CONFIG['rc'],
                'store_ai': config.STORE_CONFIG['ai']
            }
            
            for key, value in settings_to_save.items():
                # Vérifier si existe
                check = db.fetch_one("SELECT id FROM settings WHERE setting_key = ?", (key,))
                if check:
                    db.execute_update("UPDATE settings SET setting_value = ? WHERE setting_key = ?", (value, key))
                else:
                    db.execute_update("INSERT INTO settings (setting_key, setting_value) VALUES (?, ?)", (key, value))
            
            QMessageBox.information(self, "Succès", "Paramètres du magasin mis à jour!")
            
        except Exception as e:
            logger.error(f"Erreur sauvegarde store settings: {e}")
            QMessageBox.critical(self, "Erreur", f"Échec de sauvegarde: {e}")
        
    def load_users(self):
        """Charger la liste des utilisateurs"""
        try:
            # Récupérer id, username, full_name, role, et is_active
            query = "SELECT id, username, full_name, role, is_active FROM users WHERE is_active = 1"
            users = db.execute_query(query)
            
            self.users_table.setRowCount(0)
            for user in users:
                row = self.users_table.rowCount()
                self.users_table.insertRow(row)
                self.users_table.setItem(row, 0, QTableWidgetItem(user['username']))
                self.users_table.setItem(row, 1, QTableWidgetItem(user['full_name']))
                self.users_table.setItem(row, 2, QTableWidgetItem(user['role']))
                
                # Boutons d'action
                btn_widget = QWidget()
                layout = QHBoxLayout(btn_widget)
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setAlignment(Qt.AlignCenter)
                
                # Bouton Modifier mot de passe
                pwd_btn = QPushButton("🔑")
                pwd_btn.setFixedSize(30, 30)
                pwd_btn.setToolTip("Modifier le mot de passe")
                pwd_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e8f5e9; 
                        border: none; 
                        border-radius: 4px;
                        color: #27ae60;
                    }
                    QPushButton:hover {
                        background-color: #c8e6c9;
                    }
                """)
                pwd_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username']: self.change_password_dialog(uid, uname))
                
                # Bouton Permissions (Nouveau)
                perm_btn = QPushButton("🛡️")
                perm_btn.setFixedSize(30, 30)
                perm_btn.setToolTip("Gérer les permissions")
                perm_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e3f2fd; 
                        border: none; 
                        border-radius: 4px;
                        color: #1976d2;
                    }
                    QPushButton:hover {
                        background-color: #bbdefb;
                    }
                """)
                perm_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username'], urole=user['role']: self.open_permission_dialog(uid, uname, urole))

                del_btn = QPushButton("🗑️")
                del_btn.setFixedSize(30, 30)
                del_btn.setToolTip("Supprimer l'utilisateur")
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #ffcccc; 
                        border: none; 
                        border-radius: 4px;
                        color: #c0392b;
                    }
                    QPushButton:hover {
                        background-color: #ffb3b3;
                    }
                """)
                # Connecter avec l'ID et le Username
                del_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username']: self.delete_user(uid, uname))
                
                layout.addWidget(pwd_btn)
                layout.addWidget(perm_btn)
                layout.addWidget(del_btn)
                self.users_table.setCellWidget(row, 3, btn_widget)
                
        except Exception as e:
            logger.error(f"Erreur chargement utilisateurs: {e}")

    def open_permission_dialog(self, user_id, username, role):
        """Ouvrir le dialogue de permissions"""
        dialog = PermissionDialog(user_id, username, role, self)
        dialog.exec_()
        
    def change_password_dialog(self, user_id, username):
        """Ouvrir le dialogue pour changer le mot de passe"""
        from PyQt5.QtWidgets import QDialog, QFormLayout
        
        current_user = auth_manager.get_current_user()
        is_own_account = current_user and current_user['id'] == user_id
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Modifier le mot de passe - {username}")
        dialog.setMinimumWidth(350)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        # Si c'est son propre compte, demander l'ancien mot de passe
        old_pwd_input = None
        if is_own_account:
            old_pwd_input = QLineEdit()
            old_pwd_input.setEchoMode(QLineEdit.Password)
            old_pwd_input.setPlaceholderText("Votre mot de passe actuel")
            form.addRow("Ancien mot de passe:", old_pwd_input)
        
        new_pwd_input = QLineEdit()
        new_pwd_input.setEchoMode(QLineEdit.Password)
        new_pwd_input.setPlaceholderText("Minimum 4 caractères")
        
        confirm_pwd_input = QLineEdit()
        confirm_pwd_input.setEchoMode(QLineEdit.Password)
        
        form.addRow("Nouveau mot de passe:", new_pwd_input)
        form.addRow("Confirmer:", confirm_pwd_input)
        
        layout.addLayout(form)
        
        # Boutons
        btn_layout = QHBoxLayout()
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        
        save_btn = QPushButton("💾 Enregistrer")
        save_btn.setDefault(True)
        save_btn.setAutoDefault(True)
        save_btn.setStyleSheet("background-color: #27ae60; color: white;")
        
        def save_password():
            new_pwd = new_pwd_input.text()
            confirm_pwd = confirm_pwd_input.text()
            
            if new_pwd != confirm_pwd:
                QMessageBox.warning(dialog, "Erreur", "Les mots de passe ne correspondent pas")
                return
            
            if len(new_pwd) < 4:
                QMessageBox.warning(dialog, "Erreur", "Le mot de passe doit contenir au moins 4 caractères")
                return
            
            if is_own_account:
                # Utiliser change_password avec vérification de l'ancien mot de passe
                old_pwd = old_pwd_input.text()
                success, msg = auth_manager.change_password(user_id, old_pwd, new_pwd)
            else:
                # Admin peut réinitialiser sans ancien mot de passe
                from core.security import hash_password
                new_hash = hash_password(new_pwd)
                query = "UPDATE users SET password_hash = ? WHERE id = ?"
                db.execute_update(query, (new_hash, user_id))
                success, msg = True, "Mot de passe modifié avec succès"
            
            if success:
                QMessageBox.information(dialog, "Succès", msg)
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Erreur", msg)
        
        save_btn.clicked.connect(save_password)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()

    def delete_user(self, user_id, username):
        """Supprimer un utilisateur"""
        current_user = auth_manager.get_current_user()
        
        # 1. Empêcher la suppression de soi-même
        if current_user and current_user['id'] == user_id:
            QMessageBox.warning(self, "Action interdite", "Vous ne pouvez pas supprimer votre propre compte !")
            return

        reply = QMessageBox.question(self, "Confirmation", 
                                   f"Voulez-vous vraiment supprimer l'utilisateur '{username}' ?",
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            try:
                # 2. Vérifier si c'est un admin et si c'est le dernier
                all_users = auth_manager.get_all_users()
                target_user = next((u for u in all_users if u['id'] == user_id), None)
                
                if target_user and target_user['role'] == 'admin':
                    admin_count = sum(1 for u in all_users if u['role'] == 'admin')
                    if admin_count <= 1:
                        QMessageBox.warning(self, "Erreur", "Impossible de supprimer le dernier administrateur.")
                        return

                # 3. Supprimer (Soft Delete)
                query = "UPDATE users SET is_active = 0 WHERE id = ?"
                rows = db.execute_update(query, (user_id,))
                
                if rows > 0:
                    logger.info(f"Utilisateur supprimé: {username} (ID: {user_id})")
                    QMessageBox.information(self, "Succès", f"L'utilisateur {username} a été supprimé.")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Erreur", "L'utilisateur n'a pas pu être supprimé.")
                
            except Exception as e:
                logger.error(f"Erreur suppression utilisateur: {e}")
                QMessageBox.critical(self, "Erreur", f"Erreur technique: {str(e)}")
            
    def add_user(self):
        """Ajouter un utilisateur"""
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        fullname = self.fullname_input.text().strip()
        role = self.role_input.currentData()
        
        if not all([username, password, fullname]):
            QMessageBox.warning(self, "Erreur", "Tous les champs sont requis")
            return
            
        success, msg, _ = auth_manager.create_user(username, password, fullname, role)
        
        if success:
            QMessageBox.information(self, "Succès", msg)
            self.username_input.clear()
            self.password_input.clear()
            self.fullname_input.clear()
            self.load_users()
        else:
            QMessageBox.critical(self, "Erreur", msg)
            
    def toggle_dark_mode(self):
        """Basculer le mode sombre"""
        is_dark = self.dark_mode_cb.isChecked()
        self.theme_changed.emit(is_dark) # Émettre le signal pour MainWindow

    def create_tutorial_tab(self):
        """Onglet tutoriel d'utilisation"""
        from PyQt5.QtWidgets import QTextBrowser
        
        tab = QWidget()
        layout = QVBoxLayout()
        
        tutorial = QTextBrowser()
        tutorial.setOpenExternalLinks(True)
        tutorial.setHtml("""
        <h2>📖 Guide d'utilisation - Gestion Supérette AKHRIB</h2>
        
        <h3>🛒 Point de Vente (Caisse)</h3>
        <ul>
            <li><b>Scanner un produit</b> : Scannez le code-barres ou tapez-le manuellement</li>
            <li><b>Rechercher un produit</b> : Tapez le nom dans la barre de recherche</li>
            <li><b>Sélectionner un client</b> : Tapez pour rechercher ou faites défiler la liste</li>
            <li><b>Paiement à crédit</b> : Sélectionnez un client puis choisissez "Crédit"</li>
            <li><b>Raccourci</b> : Appuyez sur <b>F9</b> pour valider le paiement</li>
        </ul>
        
        <h3>📦 Gestion des Produits</h3>
        <ul>
            <li><b>Ajouter un produit</b> : Cliquez sur "➕ Nouveau Produit"</li>
            <li><b>Code-barres optionnel</b> : Si vide, un code sera généré automatiquement</li>
            <li><b>Imprimer le code-barres</b> : Cliquez sur 🏷️ dans la colonne Actions</li>
            <li><b>Importer depuis Excel</b> : Utilisez le bouton "📥 Importer"</li>
        </ul>
        
        <h3>👥 Gestion des Clients</h3>
        <ul>
            <li><b>Ajouter un client</b> : Dans la page Clients</li>
            <li><b>Crédit client</b> : Automatiquement mis à jour lors des ventes à crédit</li>
        </ul>
        
        <h3>💾 Sauvegarde des données</h3>
        <ul>
            <li><b>Exporter</b> : Paramètres → Données → Créer une Sauvegarde</li>
            <li><b>Restaurer</b> : Paramètres → Données → Restaurer depuis une Sauvegarde</li>
        </ul>
        
        <h3>🖨️ Impression</h3>
        <ul>
            <li><b>Ticket de caisse</b> : Cochez "Imprimer le ticket" avant de payer</li>
            <li><b>Codes-barres</b> : Page Produits → Bouton 🏷️</li>
        </ul>
        """)
        
        layout.addWidget(tutorial)
        tab.setLayout(layout)
        return tab

    def create_about_tab(self):
        """Onglet à propos"""
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        
        # Logo/Icône
        logo = QLabel()
        logo.setAlignment(Qt.AlignCenter)
        logo_path = str(config.LOGO_PATH)
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            scaled_pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo.setPixmap(scaled_pixmap)
        else:
            logo.setText("🏪")
            logo.setStyleSheet("font-size: 72px;")
            
        layout.addWidget(logo)
        
        # Titre
        title = QLabel("DamDev POS")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #667eea;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Version
        version = QLabel("Version 1.0.0")
        version.setStyleSheet("font-size: 14px; color: #666;")
        version.setAlignment(Qt.AlignCenter)
        layout.addWidget(version)
        
        layout.addSpacing(30)
        
        # Développeur
        dev_group = QGroupBox("👨‍💻 Développé par")
        dev_group.setStyleSheet("QGroupBox { font-size: 14px; font-weight: bold; }")
        dev_layout = QVBoxLayout()
        
        dev_name = QLabel("DamDev")
        dev_name.setStyleSheet("font-size: 20px; font-weight: bold; color: #667eea;")
        dev_name.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(dev_name)
        
        dev_layout.addSpacing(10)
        
        phone = QLabel("📞 0561491987")
        phone.setStyleSheet("font-size: 14px;")
        phone.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(phone)
        
        email = QLabel("📧 amine.drif2002@gmail.com")
        email.setStyleSheet("font-size: 14px;")
        email.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(email)
        
        dev_group.setLayout(dev_layout)
        layout.addWidget(dev_group)
        
        layout.addSpacing(20)
        
        # Copyright
        copyright_lbl = QLabel("© 2024 - Tous droits réservés")
        copyright_lbl.setStyleSheet("color: #999;")
        copyright_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(copyright_lbl)
        
        layout.addStretch()
        tab.setLayout(layout)
        return tab
    
    def set_dark_mode(self, is_dark):
        """Appliquer le mode sombre/clair"""
        if is_dark:
            # Mode sombre
            tab_style = """
                QTabWidget::pane {
                    background-color: #2c3e50;
                    border: 1px solid #4a6785;
                    border-radius: 8px;
                }
                QTabBar::tab {
                    background-color: #34495e;
                    color: #ecf0f1;
                    padding: 10px 20px;
                    border: 1px solid #4a6785;
                    border-radius: 5px 5px 0 0;
                }
                QTabBar::tab:selected {
                    background-color: #3498db;
                    color: white;
                }
            """
            group_style = """
                QGroupBox {
                    background-color: #34495e;
                    color: #ecf0f1;
                    border: 1px solid #4a6785;
                    border-radius: 10px;
                    padding: 15px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    color: #ecf0f1;
                    subcontrol-origin: margin;
                    left: 15px;
                }
            """
            table_style = """
                QTableWidget {
                    background-color: #34495e;
                    color: #ecf0f1;
                    gridline-color: #4a6785;
                    border: 1px solid #4a6785;
                    border-radius: 8px;
                }
                QTableWidget::item {
                    padding: 8px;
                    border-bottom: 1px solid #4a6785;
                }
                QTableWidget::item:selected {
                    background-color: #3498db;
                    color: white;
                }
                QTableWidget::item:alternate {
                    background-color: #2c3e50;
                }
                QHeaderView::section {
                    background-color: #3498db;
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                }
            """
            label_style = "color: #ecf0f1;"
        else:
            # Mode clair
            tab_style = """
                QTabWidget::pane {
                    background-color: #f5f5f5;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                }
                QTabBar::tab {
                    background-color: white;
                    color: #2c3e50;
                    padding: 10px 20px;
                    border: 1px solid #e0e0e0;
                    border-radius: 5px 5px 0 0;
                }
                QTabBar::tab:selected {
                    background-color: #3498db;
                    color: white;
                }
            """
            group_style = """
                QGroupBox {
                    background-color: white;
                    color: #2c3e50;
                    border: 1px solid #e0e0e0;
                    border-radius: 10px;
                    padding: 15px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    color: #2c3e50;
                    subcontrol-origin: margin;
                    left: 15px;
                }
            """
            table_style = """
                QTableWidget {
                    background-color: white;
                    color: #2c3e50;
                    gridline-color: #e0e0e0;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                }
                QTableWidget::item {
                    padding: 8px;
                    border-bottom: 1px solid #e0e0e0;
                }
                QTableWidget::item:selected {
                    background-color: #3498db;
                    color: white;
                }
                QTableWidget::item:alternate {
                    background-color: #f8f9fa;
                }
                QHeaderView::section {
                    background-color: #3498db;
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                }
            """
            label_style = "color: #2c3e50;"
        
        # Appliquer aux tabs
        if hasattr(self, 'tabs'):
            self.tabs.setStyleSheet(tab_style)
        
        # Appliquer aux tables
        if hasattr(self, 'users_table'):
            self.users_table.setStyleSheet(table_style)
        
        # Appliquer les groupes de façon récursive
        from PyQt5.QtWidgets import QGroupBox
        for group in self.findChildren(QGroupBox):
            group.setStyleSheet(group_style)

