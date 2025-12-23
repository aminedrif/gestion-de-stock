# -*- coding: utf-8 -*-
"""
Dialogue d'importation de produits
"""
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
                             QHeaderView, QMessageBox, QProgressBar)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from modules.products.product_manager import product_manager
from core.auth import auth_manager
import openpyxl

class ImportWorker(QThread):
    """Worker pour l'import en arrière-plan"""
    finished = pyqtSignal(bool, dict)
    
    def __init__(self, file_path, user_id):
        super().__init__()
        self.file_path = file_path
        self.user_id = user_id
        
    def run(self):
        success, stats = product_manager.import_products_from_excel(self.file_path, self.user_id)
        self.finished.emit(success, stats)

class ImportDialog(QDialog):
    """Dialogue d'importation de produits"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📥 Importer des Produits (Excel)")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        self.file_path = None
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Instructions
        info = QLabel("""
        <b>Instructions :</b><br>
        1. Le fichier doit être au format Excel (.xlsx)<br>
        2. Les colonnes obligatoires sont : <b>Nom</b> et <b>Prix Vente</b><br>
        3. Colonnes optionnelles : Code, Prix Achat, Stock, Min Stock
        """)
        info.setStyleSheet("background-color: #e8f6f3; padding: 10px; border-radius: 5px;")
        layout.addWidget(info)
        
        # Sélection fichier
        file_layout = QHBoxLayout()
        self.file_label = QLabel("Aucun fichier sélectionné")
        self.file_label.setStyleSheet("border: 1px solid #ccc; padding: 5px; background: white;")
        
        select_btn = QPushButton("📂 Choisir Fichier")
        select_btn.clicked.connect(self.select_file)
        
        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(select_btn)
        layout.addLayout(file_layout)
        
        # Aperçu
        layout.addWidget(QLabel("<b>Aperçu (5 premières lignes) :</b>"))
        
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(6)
        self.preview_table.setHorizontalHeaderLabels(["Code", "Nom", "Prix Achat", "Prix Vente", "Stock", "Min Stock"])
        self.preview_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        layout.addWidget(self.preview_table)
        
        # Progress Bar
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)
        
        # Boutons
        btn_layout = QHBoxLayout()
        
        template_btn = QPushButton("📋 Télécharger Modèle")
        template_btn.clicked.connect(self.download_template)
        template_btn.setStyleSheet("color: #3498db;")
        
        btn_layout.addWidget(template_btn)
        btn_layout.addStretch()
        
        self.import_btn = QPushButton("✅ Lancer l'Importation")
        self.import_btn.setEnabled(False)
        self.import_btn.clicked.connect(self.run_import)
        self.import_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 15px; font-weight: bold;")
        
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(close_btn)
        btn_layout.addWidget(self.import_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        
    def select_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Choisir fichier Excel", "", "Excel Files (*.xlsx)")
        if fname:
            self.file_path = fname
            self.file_label.setText(fname)
            self.load_preview()
            self.import_btn.setEnabled(True)
            
    def load_preview(self):
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
            
            self.preview_table.setRowCount(0)
            for row in rows:
                if not row or all(c is None for c in row): continue
                
                r = self.preview_table.rowCount()
                self.preview_table.insertRow(r)
                
                # Essayer de mapper intelligemment (ou juste afficher brut pour l'instant)
                # Amélioration possible: Utiliser la même logique de mapping que l'import
                for i, val in enumerate(row[:6]): 
                    self.preview_table.setItem(r, i, QTableWidgetItem(str(val) if val is not None else ""))
                    
        except Exception as e:
            QMessageBox.warning(self, "Erreur lecture", f"Impossible de lire le fichier : {str(e)}")
            
    def download_template(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Sauvegarder Modèle", "modele_produits.xlsx", "Excel Files (*.xlsx)")
        if fname:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Code (op)", "Nom (req)", "Prix Achat (op)", "Prix Vente (req)", "Stock (op)", "Min Stock (op)"])
                ws.append(["123456", "Exemple Produit", 100, 150, 50, 5])
                wb.save(fname)
                QMessageBox.information(self, "Succès", "Modèle sauvegardé !")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Erreur sauvegarde : {str(e)}")
                
    def run_import(self):
        if not self.file_path: return
        
        user = auth_manager.get_current_user()
        user_id = user['id'] if user else 1
        
        self.progress.setVisible(True)
        self.progress.setRange(0, 0) # Indeterminate
        self.import_btn.setEnabled(False)
        
        self.worker = ImportWorker(self.file_path, user_id)
        self.worker.finished.connect(self.on_import_finished)
        self.worker.start()
        
    def on_import_finished(self, success, stats):
        self.progress.setVisible(False)
        self.import_btn.setEnabled(True)
        
        if success:
            msg = f"""
            <b>Importation terminée !</b><br><br>
            Total lu : {stats['total']}<br>
            ✅ Succès : {stats['success']}<br>
            ⚠️ Doublons ignorés : {stats['duplicates']}<br>
            ❌ Erreurs : {stats['errors']}
            """
            if stats['errors'] > 0:
                msg += "<br><br>Détails erreurs:<br>" + "<br>".join(stats['error_details'][:5])
                
            QMessageBox.information(self, "Résultat Import", msg)
            if stats['success'] > 0:
                self.accept() # Fermer si succès
        else:
            QMessageBox.critical(self, "Erreur Import", f"Erreur critique : {stats.get('error')}")
