# -*- coding: utf-8 -*-
"""
Application de Gestion de Mini-Market
Point d'entrée principal avec interface PyQt5
"""
import sys
from pathlib import Path

# Ajouter le dossier parent au chemin Python
sys.path.append(str(Path(__file__).parent))

import config
from core.logger import logger
from database.db_manager import db

# Import PyQt5
from PyQt5.QtWidgets import QApplication, QMessageBox
from PyQt5.QtCore import Qt
from ui.login_dialog import LoginDialog
from ui.main_window import MainWindow

def initialize_application():
    """Initialiser l'application"""
    try:
        logger.info("=" * 60)
        logger.info(f"Démarrage de {config.APP_NAME} v{config.APP_VERSION}")
        logger.info("=" * 60)
        
        # Vérifier la base de données
        logger.info("Vérification de la base de données...")
        db_info = db.get_database_info()
        logger.info(f"Base de données: {db_info['path']}")
        logger.info(f"Taille: {db_info['size_bytes'] / 1024:.2f} KB")
        logger.info(f"Tables: {len(db_info['tables'])}")
        
        # Afficher les compteurs de tables
        for table, count in db_info['table_counts'].items():
            logger.info(f"  - {table}: {count} enregistrement(s)")
        
        logger.info("Application initialisée avec succès")
        return True
        
    except Exception as e:
        logger.critical(f"Erreur lors de l'initialisation: {e}")
        logger.exception("Détails de l'erreur:")
        return False

def main():
    """Fonction principale avec interface PyQt5"""
    # Initialiser l'application
    if not initialize_application():
        print("Erreur lors de l'initialisation de l'application")
        sys.exit(1)
    
    # Créer l'application Qt
    app = QApplication(sys.argv)
    app.setApplicationName(config.APP_NAME)
    app.setApplicationVersion(config.APP_VERSION)
    
    # Configurer le style
    app.setStyle('Fusion')
    
    # Afficher le dialogue de connexion
    login_dialog = LoginDialog()
    
    if login_dialog.exec_() == LoginDialog.Accepted:
        # Connexion réussie, obtenir les données utilisateur
        from core.auth import auth_manager
        user_data = auth_manager.get_current_user()
        
        if user_data:
            # Créer et afficher la fenêtre principale
            main_window = MainWindow(user_data)
            main_window.showMaximized()
            
            # Configurer la sauvegarde automatique toutes les heures
            from PyQt5.QtCore import QTimer
            from core.backup import backup_manager
            
            def perform_auto_backup():
                """Effectuer une sauvegarde automatique"""
                try:
                    # Sauvegarde SQL (base de données)
                    backup_manager.auto_backup()
                    
                    # Sauvegarde Excel
                    import openpyxl
                    from datetime import datetime
                    backup_dir = config.BACKUP_DIR
                    backup_dir.mkdir(exist_ok=True)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_path = backup_dir / f"auto_backup_{timestamp}.xlsx"
                    
                    wb = openpyxl.Workbook()
                    
                    # Produits
                    ws = wb.active
                    ws.title = "Produits"
                    products = db.execute_query("SELECT * FROM products WHERE is_active = 1")
                    if products:
                        ws.append(list(dict(products[0]).keys()))
                        for p in products:
                            ws.append(list(dict(p).values()))
                    
                    # Ventes
                    ws_sales = wb.create_sheet("Ventes")
                    sales = db.execute_query("SELECT * FROM sales")
                    if sales:
                        ws_sales.append(list(dict(sales[0]).keys()))
                        for s in sales:
                            ws_sales.append(list(dict(s).values()))
                    
                    # Clients
                    ws_cust = wb.create_sheet("Clients")
                    customers = db.execute_query("SELECT * FROM customers WHERE is_active = 1")
                    if customers:
                        ws_cust.append(list(dict(customers[0]).keys()))
                        for c in customers:
                            ws_cust.append(list(dict(c).values()))
                    
                    wb.save(excel_path)
                    logger.info(f"Sauvegarde automatique créée: {excel_path.name}")
                    
                except Exception as e:
                    logger.error(f"Erreur sauvegarde automatique: {e}")
            
            # Timer pour backup toutes les heures (3600000 ms)
            backup_timer = QTimer()
            backup_timer.timeout.connect(perform_auto_backup)
            backup_timer.start(3600000)  # 1 heure en millisecondes
            
            # Faire une première sauvegarde au démarrage
            perform_auto_backup()
            
            # Lancer la boucle d'événements
            sys.exit(app.exec_())
        else:
            logger.error("Erreur: Aucune donnée utilisateur après connexion")
            QMessageBox.critical(None, "Erreur", "Erreur lors de la récupération des données utilisateur")
            sys.exit(1)
    else:
        # Connexion annulée
        logger.info("Connexion annulée par l'utilisateur")
        sys.exit(0)

if __name__ == "__main__":
    main()
